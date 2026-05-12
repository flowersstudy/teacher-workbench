from __future__ import annotations

import json
import re
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL_DIR = Path(r"D:\xitong\other\数据导入\知识点理论课数据导入")
OUTPUT_PATH = REPO_ROOT / "src" / "pages" / "TeacherWorkbench" / "components" / "LeftPanel" / "assignmentLibrary.generated.ts"

PROVINCE_CANDIDATES = [
    "国考",
    "北京",
    "天津",
    "河北",
    "山西",
    "内蒙古",
    "辽宁",
    "吉林",
    "黑龙江",
    "上海",
    "江苏",
    "浙江",
    "安徽",
    "福建",
    "江西",
    "山东",
    "河南",
    "湖北",
    "湖南",
    "四川",
    "重庆",
    "贵州",
    "云南",
    "西藏",
    "陕西",
    "甘肃",
    "青海",
    "宁夏",
    "新疆",
    "广西",
    "海南",
    "广东",
    "深圳",
    "香港",
    "澳门",
    "台湾",
]


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", "").replace("\n", "").strip()


def normalize_header(value: Any) -> str:
    return normalize_cell(value).replace("（", "(").replace("）", ")")


def strip_trailing_empty(values: list[str]) -> list[str]:
    trimmed = values[:]
    while trimmed and trimmed[-1] == "":
        trimmed.pop()
    return trimmed


def unique_provinces(text: str) -> list[str]:
    return [province for province in PROVINCE_CANDIDATES if province and province in text]


def parse_rule(raw_text: str, *, default_type: str = "optional") -> tuple[str, str, list[str]]:
    text = normalize_cell(raw_text)
    provinces = unique_provinces(text)

    if text == "必学":
        return "required", "all", []
    if text == "选学":
        return "optional", "all", []

    if "非" in text and "必学" in text and provinces:
        return "optional", "required_except", provinces

    if "必学" in text and provinces:
        return "optional", "required_in", provinces

    if "选学" in text and provinces:
        return "optional", "only", provinces

    if provinces:
        return default_type, "only", provinces

    return default_type, "all", []


def parse_resource_rule(status_text: str, raw_title: str) -> tuple[str, str, list[str]]:
    normalized_status = normalize_cell(status_text)
    normalized_raw_title = normalize_cell(raw_title)

    if normalized_status == "必学":
        return "required", "all", []
    if normalized_status in {"选学", "选学（补考）"}:
        return "optional", "all", []

    return parse_rule(" ".join(filter(None, [normalized_status, normalized_raw_title])), default_type="optional")


def canonical_practice_slot(text: str) -> tuple[str, str]:
    compact = normalize_cell(text)
    if "刷题训练一" in compact or "刷题一" in compact:
        return "practice_1", "刷题训练一"
    if "刷题训练二" in compact or "刷题二" in compact:
        return "practice_2", "刷题训练二"
    if "刷题训练三" in compact or "刷题三" in compact:
        return "practice_3", "刷题训练三"
    return "practice_unknown", compact or "实训"


def detect_resource_bucket(area_label: str, raw_title: str, status_text: str) -> str:
    merged = " ".join(filter(None, [area_label, raw_title, status_text]))
    if "测试" in merged:
        if "二次考试" in merged or "补考" in merged:
            return "remedial"
        return "exam"
    if "二次考试" in merged or "补考" in merged:
        return "remedial"
    if "考试" in merged:
        return "exam"
    return "practice"


def detect_resource_slot(bucket: str, raw_title: str, status_text: str) -> tuple[str, str]:
    if bucket == "exam":
        return "exam", "考试"
    if bucket == "remedial":
        return "remedial", "二次考试"
    return canonical_practice_slot(" ".join(filter(None, [raw_title, status_text])))


def is_replacement_item(raw_title: str, status_text: str) -> bool:
    merged = " ".join(filter(None, [normalize_cell(raw_title), normalize_cell(status_text)]))
    return "替换" in merged


def build_header_index(header_cells: list[str]) -> dict[str, int]:
    index: dict[str, int] = {}
    for position, value in enumerate(header_cells):
        if "learning_status" not in index and "学习状态" in value:
            index["learning_status"] = position
        elif "knowledge_point" not in index and value == "知识点":
            index["knowledge_point"] = position
        elif "theory_title" not in index and "最终呈现理论课" in value:
            index["theory_title"] = position
        elif "course_status" not in index and "课程状态" in value:
            index["course_status"] = position
        elif "video_id" not in index and value == "录播课链接":
            index["video_id"] = position
        elif "pre_class_url" not in index and value == "课前作业":
            index["pre_class_url"] = position
        elif "analysis_url" not in index and value == "作业解析":
            index["analysis_url"] = position
        elif "note_text" not in index and ("备注" in value or "现有课程使用命名" in value):
            index["note_text"] = position
    if "learning_status" not in index:
        index["learning_status"] = 0
    return index


def get_value(values: list[str], index_map: dict[str, int], key: str) -> str:
    position = index_map.get(key)
    if position is None or position >= len(values):
        return ""
    return normalize_cell(values[position])


def build_theory_row(
    *,
    checkpoint_name: str,
    source_sheet: str,
    source_row: int,
    sort_order: int,
    inherited_learning_status: str,
    inherited_knowledge_point: str,
    row_values: list[str],
    header_index: dict[str, int],
) -> dict[str, Any] | None:
    learning_status = get_value(row_values, header_index, "learning_status") or inherited_learning_status
    knowledge_point = get_value(row_values, header_index, "knowledge_point") or inherited_knowledge_point
    theory_title = get_value(row_values, header_index, "theory_title")
    course_status = get_value(row_values, header_index, "course_status")
    video_id = get_value(row_values, header_index, "video_id")
    pre_class_url = get_value(row_values, header_index, "pre_class_url")
    analysis_url = get_value(row_values, header_index, "analysis_url")
    note_text = get_value(row_values, header_index, "note_text")

    has_theory_payload = any([theory_title, video_id, pre_class_url, analysis_url, note_text])
    if not has_theory_payload:
        return None

    uses_legacy_title = False
    if not theory_title and any([video_id, pre_class_url, analysis_url]) and note_text:
        theory_title = note_text
        uses_legacy_title = True

    knowledge_type, province_rule_mode, province_keys = parse_rule(learning_status, default_type="optional")

    return OrderedDict([
        ("checkpointName", checkpoint_name),
        ("knowledgePoint", knowledge_point),
        ("knowledgeType", knowledge_type),
        ("learningStatusRaw", learning_status),
        ("provinceKeys", province_keys),
        ("provinceRuleMode", province_rule_mode),
        ("courseStatus", course_status),
        ("theoryTitle", theory_title),
        ("videoId", video_id),
        ("preClassUrl", pre_class_url),
        ("analysisUrl", analysis_url),
        ("noteText", note_text),
        ("usesLegacyTitle", uses_legacy_title),
        ("sourceSheet", source_sheet),
        ("sourceRow", source_row),
        ("sortOrder", sort_order),
    ])


def build_resource_item(
    *,
    checkpoint_name: str,
    source_sheet: str,
    source_row: int,
    sort_order: int,
    area_label: str,
    status_text: str,
    raw_title: str,
    display_title: str,
    video_id: str,
    pre_class_url: str,
    analysis_url: str,
) -> tuple[str, dict[str, Any]] | None:
    bucket = detect_resource_bucket(area_label, raw_title, status_text)
    if not any([status_text, raw_title, display_title, video_id, pre_class_url, analysis_url]):
        return None

    selection_type, province_rule_mode, province_keys = parse_resource_rule(status_text, raw_title)
    slot_key, slot_label = detect_resource_slot(bucket, raw_title, status_text)
    display_name = display_title or raw_title or slot_label
    item_kind = "practice" if bucket == "practice" else "exam"
    is_replacement = is_replacement_item(raw_title, status_text)

    item = OrderedDict([
        ("id", f"{checkpoint_name}_{source_row}_{display_name}"),
        ("checkpointName", checkpoint_name),
        ("kind", item_kind),
        ("slotKey", slot_key),
        ("slotLabel", slot_label),
        ("rawTitle", raw_title),
        ("questionTitle", display_name),
        ("displayTitle", display_name),
        ("selectionStatusRaw", status_text),
        ("selectionType", selection_type),
        ("provinceKeys", province_keys),
        ("provinceRuleMode", province_rule_mode),
        ("isReplacement", is_replacement),
        ("videoId", video_id),
        ("preClassUrl", pre_class_url),
        ("analysisUrl", analysis_url),
        ("sourceSheet", source_sheet),
        ("sourceRow", source_row),
        ("sortOrder", sort_order),
    ])

    return bucket, item


def locate_excel_path() -> Path:
    if len(sys.argv) > 1:
        path = Path(sys.argv[1]).expanduser()
        if not path.exists():
            raise FileNotFoundError(f"Excel not found: {path}")
        return path

    if not DEFAULT_EXCEL_DIR.exists():
        raise FileNotFoundError(f"Default Excel directory not found: {DEFAULT_EXCEL_DIR}")

    matches = sorted(DEFAULT_EXCEL_DIR.glob("*.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError(f"No .xlsx files found in: {DEFAULT_EXCEL_DIR}")
    return matches[0]


def generate_library(excel_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(excel_path, read_only=False, data_only=True)
    checkpoints: list[dict[str, Any]] = []

    for sheet in workbook.worksheets:
        checkpoint_name = normalize_cell(sheet.title)
        header_row = [normalize_header(sheet.cell(row=1, column=column).value) for column in range(1, sheet.max_column + 1)]
        trimmed_header = strip_trailing_empty(header_row)
        header_index = build_header_index(trimmed_header)

        if len(trimmed_header) < 7:
            raise ValueError(f"Sheet header is too short: {checkpoint_name}")

        right_start = len(trimmed_header) - 7

        theory_rows: list[dict[str, Any]] = []
        practice_items: list[dict[str, Any]] = []
        exam_items: list[dict[str, Any]] = []
        remedial_items: list[dict[str, Any]] = []

        inherited_learning_status = ""
        inherited_knowledge_point = ""
        inherited_area_label = ""
        theory_sort = 0
        resource_sort = 0

        for source_row in range(2, sheet.max_row + 1):
            row_values = [normalize_cell(sheet.cell(row=source_row, column=column).value) for column in range(1, sheet.max_column + 1)]

            current_learning_status = get_value(row_values, header_index, "learning_status")
            current_knowledge_point = get_value(row_values, header_index, "knowledge_point")
            if current_learning_status:
                inherited_learning_status = current_learning_status
            if current_knowledge_point:
                inherited_knowledge_point = current_knowledge_point

            theory_row = build_theory_row(
                checkpoint_name=checkpoint_name,
                source_sheet=checkpoint_name,
                source_row=source_row,
                sort_order=theory_sort + 1,
                inherited_learning_status=inherited_learning_status,
                inherited_knowledge_point=inherited_knowledge_point,
                row_values=row_values,
                header_index=header_index,
            )
            if theory_row is not None:
                theory_sort += 1
                theory_row["sortOrder"] = theory_sort
                theory_rows.append(theory_row)

            right_values = row_values[right_start:right_start + 7]
            right_values += [""] * max(0, 7 - len(right_values))

            area_label = normalize_cell(right_values[0]) or inherited_area_label
            if normalize_cell(right_values[0]):
                inherited_area_label = area_label

            resource_payload = build_resource_item(
                checkpoint_name=checkpoint_name,
                source_sheet=checkpoint_name,
                source_row=source_row,
                sort_order=resource_sort + 1,
                area_label=area_label,
                status_text=normalize_cell(right_values[1]),
                raw_title=normalize_cell(right_values[2]),
                display_title=normalize_cell(right_values[3]),
                video_id=normalize_cell(right_values[4]),
                pre_class_url=normalize_cell(right_values[5]),
                analysis_url=normalize_cell(right_values[6]),
            )

            if resource_payload is not None:
                bucket, item = resource_payload
                resource_sort += 1
                item["sortOrder"] = resource_sort
                if bucket == "practice":
                    practice_items.append(item)
                elif bucket == "exam":
                    exam_items.append(item)
                else:
                    remedial_items.append(item)

        checkpoints.append(OrderedDict([
            ("checkpointName", checkpoint_name),
            ("theoryRows", theory_rows),
            ("practiceItems", practice_items),
            ("examItems", exam_items),
            ("remedialItems", remedial_items),
        ]))

    return checkpoints


def render_typescript(data: list[dict[str, Any]]) -> str:
    json_payload = json.dumps(data, ensure_ascii=False, indent=2)
    return (
        "export type AssignmentKnowledgeType = 'required' | 'optional'\n"
        "export type AssignmentResourceKind = 'practice' | 'exam'\n"
        "export type AssignmentProvinceRuleMode = 'all' | 'only' | 'required_in' | 'required_except'\n\n"
        "export interface AssignmentTheoryRow {\n"
        "  checkpointName: string\n"
        "  knowledgePoint: string\n"
        "  knowledgeType: AssignmentKnowledgeType\n"
        "  learningStatusRaw: string\n"
        "  provinceKeys: string[]\n"
        "  provinceRuleMode: AssignmentProvinceRuleMode\n"
        "  courseStatus: string\n"
        "  theoryTitle: string\n"
        "  videoId: string\n"
        "  preClassUrl: string\n"
        "  analysisUrl: string\n"
        "  noteText: string\n"
        "  usesLegacyTitle: boolean\n"
        "  sourceSheet: string\n"
        "  sourceRow: number\n"
        "  sortOrder: number\n"
        "}\n\n"
        "export interface AssignmentResourceItem {\n"
        "  id: string\n"
        "  checkpointName: string\n"
        "  kind: AssignmentResourceKind\n"
        "  slotKey: string\n"
        "  slotLabel: string\n"
        "  rawTitle: string\n"
        "  questionTitle: string\n"
        "  displayTitle: string\n"
        "  selectionStatusRaw: string\n"
        "  selectionType: AssignmentKnowledgeType\n"
        "  provinceKeys: string[]\n"
        "  provinceRuleMode: AssignmentProvinceRuleMode\n"
        "  isReplacement: boolean\n"
        "  videoId: string\n"
        "  preClassUrl: string\n"
        "  analysisUrl: string\n"
        "  sourceSheet: string\n"
        "  sourceRow: number\n"
        "  sortOrder: number\n"
        "}\n\n"
        "export interface AssignmentCheckpointLibrary {\n"
        "  checkpointName: string\n"
        "  theoryRows: AssignmentTheoryRow[]\n"
        "  practiceItems: AssignmentResourceItem[]\n"
        "  examItems: AssignmentResourceItem[]\n"
        "  remedialItems: AssignmentResourceItem[]\n"
        "}\n\n"
        f"export const CHECKPOINT_ASSIGNMENT_LIBRARY: AssignmentCheckpointLibrary[] = {json_payload}\n"
    )


def main() -> None:
    excel_path = locate_excel_path()
    library = generate_library(excel_path)
    OUTPUT_PATH.write_text(render_typescript(library), encoding="utf-8")
    print(f"Generated {OUTPUT_PATH} from {excel_path}")


if __name__ == "__main__":
    main()
